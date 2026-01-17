#!/usr/bin/env python3
"""
Surface Tension Tester - Full Python Control
Integrated with Arduino v5.0 Mode-Based Architecture
"""

import serial
import serial.tools.list_ports
import json
import time
import os
from collections import defaultdict
from datetime import datetime
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading

class SurfaceTensionController:
    def __init__(self, port='COM5', baudrate=115200):
        """Initialize the controller"""
        self.port = port
        self.baudrate = baudrate
        self.ser = None
        self.running = False
        
        # Test profiles (must match Arduino)
        self.tests = {
            '1': {'name': 'ULTRA_FAST', 'speed': 1600, 'speed_mms': 1.000, 'desc': 'Ultra Fast (1 mm/s)'},
            '2': {'name': 'FAST_UP', 'speed': 1200, 'speed_mms': 0.750, 'desc': 'Fast Up (750 µm/s)'},
            '3': {'name': 'FAST_DOWN', 'speed': 400, 'speed_mms': 0.250, 'desc': 'Fast Down (250 µm/s)'},
            '4': {'name': 'MEASURE_F', 'speed': 50.0, 'speed_mms': 0.03125, 'desc': 'Measure Fast (31.25 µm/s)'},
            '5': {'name': 'MEASURE_M', 'speed': 20.0, 'speed_mms': 0.0125, 'desc': 'Measure Medium (12.5 µm/s)'},
            '6': {'name': 'MEASURE_U', 'speed': 10.0, 'speed_mms': 0.00625, 'desc': 'Measure Ultra (6.25 µm/s)'},
            '7': {'name': 'MEASURE_X', 'speed': 5.0, 'speed_mms': 0.003125, 'desc': 'Measure Extra (3.125 µm/s)'},
            '8': {'name': 'MEASURE_Z', 'speed': 2.0, 'speed_mms': 0.00125, 'desc': 'Measure Zero (1.25 µm/s)'},
        }
        
        # Data storage
        self.all_test_data = {}  # {speed_name: {'runs': [{'times': [], 'forces': [], 'positions': []}], 'peak_force': [], ...}}
        self.current_test_data = None
        self.current_test_info = None
        self.current_test_index = -1
        
        # Output directory
        self.output_dir = r"C:\Users\srnan\Documents\pythontest"
        self.ensure_output_dir()
        
        # Connection
        self.connect_serial()
        
    def ensure_output_dir(self):
        """Ensure output directory exists"""
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            print(f"✓ Output directory: {self.output_dir}")
        except Exception as e:
            print(f"⚠ Warning: Could not create directory: {e}")
            self.output_dir = os.getcwd()
    
    def connect_serial(self):
        """Connect to Arduino"""
        try:
            self.ser = serial.Serial(self.port, self.baudrate, timeout=1)
            time.sleep(2)  # Wait for Arduino to reset
            print(f"✓ Connected to {self.port} at {self.baudrate} baud")
            return True
        except Exception as e:
            print(f"✗ Error connecting to {self.port}: {e}")
            self.list_ports()
            return False
    
    def list_ports(self):
        """List available ports"""
        print("\nAvailable ports:")
        ports = [port.device for port in serial.tools.list_ports.comports()]
        if ports:
            for port in ports:
                print(f"  - {port}")
        else:
            print("  No ports found")
    
    def send_command(self, cmd):
        """Send command to Arduino"""
        try:
            if self.ser and self.ser.is_open:
                self.ser.write(cmd.encode() + b'\n')
                time.sleep(0.1)
                return True
        except Exception as e:
            print(f"✗ Error sending command: {e}")
        return False
    
    def read_data_stream(self, duration=60):
        """Read data stream from Arduino for a duration"""
        print(f"\n⏱ Reading data for {duration} seconds...")
        self.current_test_data = {'times': [], 'forces': [], 'positions': []}
        
        start_time = time.time()
        last_print = start_time
        data_count = 0
        stream_ended = False
        home_reached = False
        
        try:
            while (time.time() - start_time) < duration and self.ser.is_open:
                if self.ser.in_waiting:
                    line = self.ser.readline().decode('utf-8', errors='ignore').strip()
                    
                    if not line:
                        continue
                    
                    # Parse JSON data
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            data = json.loads(line)
                            time_val = data.get('t', 0)
                            force = data.get('f', 0)
                            position = data.get('p', 0)
                            
                            self.current_test_data['times'].append(time_val)
                            self.current_test_data['forces'].append(force)
                            self.current_test_data['positions'].append(position)
                            data_count += 1
                            
                            # Print progress every 5 seconds
                            if (time.time() - last_print) >= 5:
                                elapsed = time.time() - start_time
                                print(f"  {elapsed:.1f}s: {data_count} data points, Force: {force:.5f}N")
                                last_print = time.time()
                        except json.JSONDecodeError:
                            pass
                    
                    # Check for end marker or home reached
                    elif "END_STREAM" in line:
                        stream_ended = True
                        print(f"  Stream ended at {data_count} points")
                        # Continue waiting for HOME confirmation
                    
                    elif "HOME reached" in line:
                        home_reached = True
                        print("✓ Motor returned to HOME position")
                        break
                    
                    # Suppress other messages
                    elif line and not line.startswith("="):
                        pass
                
                time.sleep(0.01)
        
        except KeyboardInterrupt:
            print("\n✗ Data reading interrupted")
        
        print(f"✓ Data reading complete: {data_count} points collected")
        return len(self.current_test_data['times']) > 0
    
    def calculate_peak_force(self):
        """Calculate peak force from current test data"""
        if not self.current_test_data or not self.current_test_data['forces']:
            return 0
        return max(self.current_test_data['forces'])
    
    def save_to_excel(self):
        """Save all test data to Excel file with statistics"""
        if not self.all_test_data:
            print("✗ No data to save")
            return False
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(self.output_dir, f"surface_tension_results_{timestamp}.xlsx")
        
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Create sheets for each test
            for test_name, test_info in self.all_test_data.items():
                runs = test_info['runs']
                peak_forces = test_info['peak_forces']
                speed_mms = test_info['speed_mms']
                num_runs = len(runs)
                
                # Create summary sheet
                ws_summary = wb.create_sheet(title=f"{test_name}_Summary")
                
                # Header
                ws_summary['A1'] = f"Test Summary - {test_name}"
                ws_summary['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws_summary['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                ws_summary.merge_cells('A1:D1')
                ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Test info
                ws_summary['A3'] = "Test Information:"
                ws_summary['A3'].font = Font(bold=True, size=11)
                
                ws_summary['A4'] = "Speed (µm/s):"
                ws_summary['B4'] = speed_mms * 1000
                ws_summary['A5'] = "Number of Runs:"
                ws_summary['B5'] = num_runs
                
                # Statistics
                if num_runs > 0:
                    avg_peak = np.mean(peak_forces)
                    std_peak = np.std(peak_forces)
                    rsd = (std_peak / avg_peak * 100) if avg_peak != 0 else 0
                    
                    ws_summary['A7'] = "Peak Force Statistics:"
                    ws_summary['A7'].font = Font(bold=True, size=11)
                    
                    ws_summary['A8'] = "Average Peak Force (N):"
                    ws_summary['B8'] = round(avg_peak, 6)
                    ws_summary['A9'] = "Std Dev (N):"
                    ws_summary['B9'] = round(std_peak, 6)
                    ws_summary['A10'] = "RSD (%):"
                    ws_summary['B10'] = round(rsd, 2)
                    ws_summary['A11'] = "Min Peak (N):"
                    ws_summary['B11'] = round(np.min(peak_forces), 6)
                    ws_summary['A12'] = "Max Peak (N):"
                    ws_summary['B12'] = round(np.max(peak_forces), 6)
                
                # Peak forces table
                ws_summary['A14'] = "Individual Runs:"
                ws_summary['A14'].font = Font(bold=True, size=11)
                
                ws_summary['A15'] = "Run #"
                ws_summary['B15'] = "Peak Force (N)"
                ws_summary['C15'] = "Data Points"
                ws_summary['D15'] = "Duration (s)"
                
                header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                header_font = Font(bold=True)
                for col in ['A', 'B', 'C', 'D']:
                    ws_summary[f'{col}15'].fill = header_fill
                    ws_summary[f'{col}15'].font = header_font
                    ws_summary[f'{col}15'].alignment = Alignment(horizontal='center')
                
                # Individual run data
                for run_idx, (run_data, peak_force) in enumerate(zip(runs, peak_forces), 1):
                    row = 15 + run_idx
                    ws_summary[f'A{row}'] = run_idx
                    ws_summary[f'B{row}'] = round(peak_force, 6)
                    ws_summary[f'C{row}'] = len(run_data['times'])
                    ws_summary[f'D{row}'] = round(run_data['times'][-1], 2) if run_data['times'] else 0
                    
                    for col in ['A', 'B', 'C', 'D']:
                        ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center')
                
                # Auto width
                ws_summary.column_dimensions['A'].width = 25
                ws_summary.column_dimensions['B'].width = 18
                ws_summary.column_dimensions['C'].width = 15
                ws_summary.column_dimensions['D'].width = 15
                
                # Create detailed data sheets for each run
                for run_idx, run_data in enumerate(runs, 1):
                    ws_data = wb.create_sheet(title=f"{test_name}_R{run_idx}")
                    
                    # Header
                    ws_data['A1'] = f"{test_name} - Run #{run_idx}"
                    ws_data['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                    ws_data['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    ws_data.merge_cells('A1:D1')
                    ws_data['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Run info
                    ws_data['A3'] = "Run Information:"
                    ws_data['A3'].font = Font(bold=True, size=11)
                    
                    ws_data['A4'] = "Speed (µm/s):"
                    ws_data['B4'] = speed_mms * 1000
                    ws_data['A5'] = "Peak Force (N):"
                    ws_data['B5'] = round(peak_forces[run_idx - 1], 6)
                    ws_data['A6'] = "Duration (s):"
                    ws_data['B6'] = round(run_data['times'][-1], 2) if run_data['times'] else 0
                    ws_data['A7'] = "Data Points:"
                    ws_data['B7'] = len(run_data['times'])
                    
                    # Data table header
                    ws_data['A9'] = "Time (s)"
                    ws_data['B9'] = "Force (N)"
                    ws_data['C9'] = "Position (mm)"
                    
                    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    header_font = Font(bold=True)
                    for col in ['A', 'B', 'C']:
                        ws_data[f'{col}9'].fill = header_fill
                        ws_data[f'{col}9'].font = header_font
                        ws_data[f'{col}9'].alignment = Alignment(horizontal='center')
                    
                    # Data rows
                    for row_idx, (time, force, pos) in enumerate(zip(run_data['times'], run_data['forces'], run_data['positions']), start=10):
                        ws_data[f'A{row_idx}'] = round(time, 3)
                        ws_data[f'B{row_idx}'] = round(force, 5)
                        ws_data[f'C{row_idx}'] = round(pos, 4)
                        
                        for col in ['A', 'B', 'C']:
                            ws_data[f'{col}{row_idx}'].alignment = Alignment(horizontal='center')
                    
                    # Border
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    for row in ws_data.iter_rows(min_row=9, max_row=len(run_data['times'])+9, min_col=1, max_col=3):
                        for cell in row:
                            cell.border = thin_border
                    
                    # Auto width
                    ws_data.column_dimensions['A'].width = 15
                    ws_data.column_dimensions['B'].width = 15
                    ws_data.column_dimensions['C'].width = 15
            
            wb.save(filename)
            print(f"✓ Data saved to: {filename}")
            return filename
        
        except Exception as e:
            print(f"✗ Error saving Excel file: {e}")
            return False
    
    def plot_results(self):
        """
        Plot all test results: Force (N) vs Displacement (mm)
        DESIGN CHOICE: Grayscale + subtle color accents (hybrid approach)
        
        Rationale:
        - Primary data (curves): Grayscale with distinct line styles
          → Survives monochrome printing without loss of clarity
          → Meets strict journal standards (e.g., Nature, Science, ACS)
          → Line styles (solid, dashed) ensure accessibility
        
        - Peak force annotations: Subtle muted color (single color per test speed)
          → Adds visual hierarchy without compromising B&W reproduction
          → Helps readers on screen distinguish measurement uncertainties
          → Color is functional (links annotation to curve), not decorative
        
        - Rationale for hybrid over pure B&W:
          • Pure B&W would require many line styles → visual complexity
          • Multiple runs per speed need distinction → color helps
          • Peak annotations benefit from color coding → improves clarity
          • Hybrid respects printability while aiding digital readership
        
        - Rationale for hybrid over full color:
          • Full pastel colors are unnecessary for B&W printable figures
          • Adds costs for color journals; limits submission venues
          • Grayscale data is more neutral and objective
          • Color reserved only for annotations (functional use)
        """
        if not self.all_test_data:
            print("✗ No data to plot")
            return
        
        # ========== STEP 1: Sort tests by speed (highest to lowest) ==========
        sorted_tests = sorted(
            self.all_test_data.items(),
            key=lambda x: x[1]['speed_mms'],
            reverse=True
        )
        num_tests = len(sorted_tests)
        
        # ========== STEP 2: Configure matplotlib for journal publication ==========
        plt.rcParams['font.family'] = 'serif'
        plt.rcParams['font.serif'] = ['Times New Roman', 'DejaVu Serif']
        plt.rcParams['font.size'] = 10
        plt.rcParams['axes.linewidth'] = 0.8
        plt.rcParams['xtick.major.width'] = 0.8
        plt.rcParams['ytick.major.width'] = 0.8
        
        # Create figure with stacked subplots (one per speed)
        fig, axes = plt.subplots(
            num_tests, 1,
            figsize=(7.5, 2.5 * num_tests),
            sharex=False,
            sharey=True
        )
        
        if num_tests == 1:
            axes = [axes]
        
        fig.patch.set_facecolor('white')
        
        # ========== STEP 3: Define color palette for annotations only ==========
        # Subtle, muted accent colors used ONLY for peak annotations
        # These colors aid on-screen readability without compromising B&W printing
        accent_colors = [
            '#6B8E99',  # Muted steel blue
            '#7A8F6B',  # Muted sage
            '#8B7A99',  # Muted dusty purple
            '#996B6B',  # Muted warm brown
            '#6B9999',  # Muted slate
            '#999B6B',  # Muted olive
        ]
        
        # ========== STEP 4: Define grayscale and line styles for data curves ==========
        # Grayscale: ensures B&W printability without loss of information
        # Line styles: solid, dashed, dotted → distinguish multiple runs
        grayscale_colors = ['#000000', '#404040', '#707070', '#A0A0A0']
        line_styles = ['-', '--', '-.', ':']
        
        # Track global y-axis limits for consistency
        y_min_global = float('inf')
        y_max_global = float('-inf')
        
        # First pass: calculate global y-axis range
        for test_name, test_info in sorted_tests:
            runs = test_info['runs']
            for run_data in runs:
                forces = np.array(run_data['forces'])
                if len(forces) > 0:
                    y_min_global = min(y_min_global, np.min(forces))
                    y_max_global = max(y_max_global, np.max(forces))
        
        # Add 10% padding to y-axis
        y_range = y_max_global - y_min_global
        y_min_global -= 0.1 * y_range
        y_max_global += 0.1 * y_range
        
        # ========== Main plotting loop ==========
        for idx, (test_name, test_info) in enumerate(sorted_tests):
            ax = axes[idx]
            
            runs = test_info['runs']
            peak_forces = test_info['peak_forces']
            speed_mms = test_info['speed_mms']
            num_runs = len(runs)
            
            # Get accent color for this test speed (used for annotations only)
            accent_color = accent_colors[idx % len(accent_colors)]
            
            # Plot each run as Force vs Displacement
            for run_idx, run_data in enumerate(runs):
                times = np.array(run_data['times'])
                forces = np.array(run_data['forces'])
                positions = np.array(run_data['positions'])
                
                # ========== STEP 5: Trim trailing zero-force data ==========
                if len(forces) > 0 and np.max(forces) > 0:
                    threshold = np.max(forces) * 0.001
                    last_significant = np.where(forces > threshold)[0]
                    if len(last_significant) > 0:
                        cutoff_idx = last_significant[-1] + int(0.05 * len(forces))
                        cutoff_idx = min(cutoff_idx, len(forces) - 1)
                        forces_trimmed = forces[:cutoff_idx]
                        positions_trimmed = positions[:cutoff_idx]
                    else:
                        forces_trimmed = forces
                        positions_trimmed = positions
                else:
                    forces_trimmed = forces
                    positions_trimmed = positions
                
                # ========== STEP 6: Plot data curves in grayscale with line style variation ==========
                # Grayscale ensures the figure prints clearly in monochrome
                # Line styles (solid, dashed) distinguish multiple runs without color
                gray_color = grayscale_colors[run_idx % len(grayscale_colors)]
                line_style = line_styles[run_idx % len(line_styles)]
                label = f"Run {run_idx + 1}"
                
                ax.plot(
                    positions_trimmed, forces_trimmed,
                    linestyle=line_style,
                    color=gray_color,
                    linewidth=1.8,
                    label=label,
                    alpha=0.85,
                    zorder=2
                )
                
                # ========== STEP 7: Annotate peak force with functional color accent ==========
                # Color is used here as a FUNCTIONAL element (not decorative)
                # It visually links the annotation to the parent curve
                # This aids on-screen readability without compromising B&W printing
                peak_idx = np.argmax(forces_trimmed) if len(forces_trimmed) > 0 else 0
                if len(positions_trimmed) > peak_idx:
                    peak_pos = positions_trimmed[peak_idx]
                    peak_force = forces_trimmed[peak_idx]
                    
                    # Thin horizontal dashed line at peak force level (grayscale)
                    ax.axhline(
                        peak_force,
                        color=gray_color,
                        linestyle=':',
                        linewidth=0.8,
                        alpha=0.3,
                        zorder=1
                    )
                    
                    # Thin vertical dashed line at peak position (grayscale)
                    ax.axvline(
                        peak_pos,
                        color=gray_color,
                        linestyle=':',
                        linewidth=0.8,
                        alpha=0.3,
                        zorder=1
                    )
                    
                    # Small marker at peak (uses accent color for visual link)
                    ax.plot(
                        peak_pos, peak_force,
                        'o',
                        color=accent_color,
                        markersize=6,
                        markeredgecolor=gray_color,
                        markeredgewidth=1.0,
                        zorder=3
                    )
                    
                    # ========== STEP 8: Annotate peak force value ==========
                    # Text color matches accent color (functional link to curve)
                    # Annotation is minimal and positioned to avoid clutter
                    peak_text = f"Peak = {peak_force:.4f} N"
                    
                    # Calculate offset for text position
                    y_offset = 0.05 * (y_max_global - y_min_global)
                    x_offset = 0.02 * (ax.get_xlim()[1] - ax.get_xlim()[0])
                    
                    ax.annotate(
                        peak_text,
                        xy=(peak_pos, peak_force),
                        xytext=(peak_pos + x_offset, peak_force + y_offset),
                        fontsize=8,
                        color=accent_color,
                        fontweight='normal',
                        zorder=4,
                        bbox=dict(
                            boxstyle='round,pad=0.3',
                            facecolor='white',
                            edgecolor='none',
                            alpha=0.8
                        ),
                        ha='left',
                        va='bottom'
                    )
            
            # ========== STEP 9: Axis labels ==========
            ax.set_xlabel('Displacement (mm)', fontsize=11, fontweight='normal')
            ax.set_ylabel('Force (N)', fontsize=11, fontweight='normal')
            
            # ========== STEP 10: Subplot titles (minimal, informative) ==========
            speed_label = f"{speed_mms * 1000:.2f} µm/s"
            title = f"{test_name} — {speed_label}"
            ax.set_title(
                title,
                fontsize=10,
                fontweight='bold',
                loc='left',
                pad=8
            )
            
            # ========== STEP 11: Grid (light, reference only) ==========
            ax.grid(True, alpha=0.2, linestyle='-', linewidth=0.5, color='gray')
            ax.set_axisbelow(True)
            
            # ========== STEP 12: Consistent y-axis limits ==========
            ax.set_ylim([y_min_global, y_max_global])
            
            # ========== STEP 13: Legend (clean, informative) ==========
            if num_runs > 1:
                ax.legend(
                    loc='upper left',
                    fontsize=9,
                    framealpha=0.95,
                    edgecolor='black',
                    fancybox=False,
                    ncol=1
                )
            
            # ========== STEP 14: Minimal tick parameters ==========
            ax.tick_params(axis='both', which='major', labelsize=9)
            ax.set_facecolor('white')
            
            # Remove decorative spines
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
        
        # ========== STEP 15: Figure-level title ==========
        fig.suptitle(
            'Surface Tension Measurements',
            fontsize=12,
            fontweight='bold',
            y=0.998
        )
        
        # ========== STEP 16: Layout and output ==========
        plt.tight_layout()
        
        # Save figure for journal submission
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(self.output_dir, f"surface_tension_plot_{timestamp}.pdf")
        try:
            plt.savefig(output_path, dpi=300, bbox_inches='tight', format='pdf')
            print(f"✓ Plot saved to: {output_path}")
        except Exception as e:
            print(f"⚠ Could not save PDF: {e}")
        
        plt.show()
    
    def show_menu(self):
        """Display main menu"""
        print("\n" + "="*70)
        print("SURFACE TENSION TESTER v5.0 - Python Control".center(70))
        print("="*70)
        print("\nMODE COMMANDS:")
        print("  M     - LIVE MONITOR (display force continuously)")
        print("  1-8   - NORMAL TEST (single speed)")
        print("  A     - AUTO SEQUENCE (all 8 speeds)")
        print("  C     - CALIBRATION (calibrate load cell)")
        print("\nUTILITIES:")
        print("  T     - Tare scale (zero force)")
        print("  0     - Set home position")
        print("  H     - Go home")
        print("\nCONTROL:")
        print("  Q     - Quit without saving")
        print("  S     - Save and display results, then exit")
        print("="*70)
        print("\nSpeed Options:")
        for key, test in self.tests.items():
            print(f"  {key}. {test['desc']}")
        print("="*70)
    
    def run(self):
        """Main control loop"""
        if not self.ser or not self.ser.is_open:
            print("✗ Failed to connect to Arduino")
            return
        
        print("\n✓ System Ready! Waiting for Arduino initialization...")
        time.sleep(2)
        
        # Read initial messages from Arduino
        print("\nInitializing Arduino...")
        init_time = time.time()
        while (time.time() - init_time) < 3:
            if self.ser.in_waiting:
                line = self.ser.readline().decode('utf-8', errors='ignore').strip()
                if "System Ready" in line or "✓" in line:
                    print(f"  {line}")
            time.sleep(0.1)
        
        while True:
            self.show_menu()
            choice = input("\nEnter command (M/1-8/A/C/T/0/H/Q/S): ").strip().upper()
            
            if choice == 'Q':
                print("\n✗ Exiting without saving...")
                break
            
            elif choice == 'S':
                print("\nSaving and displaying results...")
                excel_file = self.save_to_excel()
                if excel_file:
                    self.plot_results()
                break
            
            elif choice == 'M':
                print("\n>>> Entering LIVE MONITOR mode...")
                self.send_command('m')
                print("Force display streaming from Arduino.")
                print("Press Ctrl+C to stop and return to menu.\n")
                
                try:
                    while True:
                        if self.ser.in_waiting:
                            line = self.ser.readline().decode('utf-8', errors='ignore').strip()
                            
                            # Look for force data from Arduino
                            if "Force:" in line:
                                print(line, end='\r')  # Overwrite on same line
                        
                        time.sleep(0.05)
                except KeyboardInterrupt:
                    self.send_command('q')
                    print("\n\nReturning to main menu...")
                    time.sleep(1)
            
            elif choice == 'A':
                print("\n🚀 Starting AUTO SEQUENCE (all 8 speeds)...")
                self.send_command('a')
                
                # Read all auto sequence data
                test_count = 0
                current_test = None
                
                try:
                    while test_count < 8:
                        if self.ser.in_waiting:
                            line = self.ser.readline().decode('utf-8', errors='ignore').strip()
                            
                            if "TEST" in line and "/" in line:
                                print(f"\n{line}")
                                # Extract test name from Arduino output
                                current_test = line.split(":")[-1].strip() if ":" in line else None
                            
                            elif line.startswith('{') and line.endswith('}'):
                                try:
                                    data = json.loads(line)
                                    if self.current_test_data is None:
                                        self.current_test_data = {'times': [], 'forces': [], 'positions': []}
                                    
                                    self.current_test_data['times'].append(data.get('t', 0))
                                    self.current_test_data['forces'].append(data.get('f', 0))
                                    self.current_test_data['positions'].append(data.get('p', 0))
                                except json.JSONDecodeError:
                                    pass
                            
                            elif "END_STREAM" in line:
                                if self.current_test_data and current_test:
                                    peak = max(self.current_test_data['forces']) if self.current_test_data['forces'] else 0
                                    self.all_test_data[current_test] = {
                                        'times': self.current_test_data['times'],
                                        'forces': self.current_test_data['forces'],
                                        'positions': self.current_test_data['positions'],
                                        'speed_mms': self.tests[str(test_count + 1)]['speed_mms'],
                                        'peak_force': peak,
                                    }
                                    test_count += 1
                                    self.current_test_data = None
                                    print(f"  ✓ Data collected (Peak: {peak:.5f}N)")
                            
                            elif "AUTO SEQUENCE COMPLETE" in line:
                                print(f"\n{line}")
                                test_count = 8  # Exit loop
                            
                            elif line and not line.startswith("="):
                                print(f"  {line}")
                        
                        time.sleep(0.01)
                except KeyboardInterrupt:
                    self.send_command('q')
                    print("\nAuto sequence interrupted.")
                    time.sleep(1)
            
            elif choice in self.tests:
                self.run_single_test(choice)
            
            elif choice == 'T':
                print("\n>>> Sending Tare command to Arduino...")
                self.send_command('t')
                time.sleep(2)
                print("✓ Tare complete")
            
            elif choice == '0':
                print("\n>>> Setting HOME position...")
                self.send_command('0')
                time.sleep(1)
                print("✓ Home position set")
            
            elif choice == 'H':
                print("\n>>> Going HOME...")
                self.send_command('h')
                time.sleep(3)
                print("✓ Home reached")
            
            elif choice == 'C':
                print("\n>>> Entering CALIBRATION mode...")
                self.send_command('c')
                print("Follow on-screen calibration steps.")
                # Let Arduino handle calibration, Python just passes through
                try:
                    while True:
                        if self.ser.in_waiting:
                            line = self.ser.readline().decode('utf-8', errors='ignore').strip()
                            print(line)
                            
                            if "Exiting calibration" in line or "ready for normal measurement" in line:
                                break
                        time.sleep(0.1)
                except KeyboardInterrupt:
                    pass
                time.sleep(1)
            
            else:
                print("✗ Invalid command")
    
    def run_single_test(self, test_num):
        """Run a single test with repeat capability"""
        if test_num not in self.tests:
            print("✗ Invalid test number")
            return
        
        test_info = self.tests[test_num]
        test_name = test_info['name']
        speed_mms = test_info['speed_mms']
        
        # Check if test already exists
        if test_name in self.all_test_data:
            print(f"\n⚠️  Test '{test_name}' already has data!")
            num_runs = len(self.all_test_data[test_name]['runs'])
            print(f"   Current runs: {num_runs}")
            
            repeat = input("   Repeat test? (y/n): ").strip().lower()
            if repeat != 'y':
                return
            
            print(f"\n   Running test #{num_runs + 1} for {test_name}...")
        else:
            print(f"\n{'='*70}")
            print(f"TEST {test_num}: {test_info['desc']}")
            print(f"{'='*70}")
            
            # Initialize test data structure for first run
            self.all_test_data[test_name] = {
                'runs': [],
                'speed_mms': speed_mms,
                'peak_forces': [],
            }
        
        # Send command to Arduino
        self.send_command(test_num)
        time.sleep(1)
        
        # Estimate test duration
        estimated_duration = (15.0 / speed_mms) + 5
        estimated_duration = int(estimated_duration) + 20  # Extended buffer
        
        print(f"Speed: {speed_mms*1000:.3f} µm/s")
        print(f"Estimated duration: ~{estimated_duration}s (includes HOME return)")
        
        # Read data
        if self.read_data_stream(duration=estimated_duration):
            peak_force = self.calculate_peak_force()
            
            # Validate data (check if peak force is reasonable)
            if peak_force < 0.00001 and len(self.current_test_data['times']) > 100:
                print("\n⚠️  WARNING: Peak force seems too low!")
                print(f"   Peak: {peak_force:.5f}N with {len(self.current_test_data['times'])} points")
                retry = input("   Retry this test? (y/n): ").strip().lower()
                if retry == 'y':
                    time.sleep(1)
                    self.run_single_test(test_num)
                    return
            
            # Store this run
            run_data = {
                'times': self.current_test_data['times'],
                'forces': self.current_test_data['forces'],
                'positions': self.current_test_data['positions'],
            }
            
            self.all_test_data[test_name]['runs'].append(run_data)
            self.all_test_data[test_name]['peak_forces'].append(peak_force)
            
            # Calculate statistics
            num_runs = len(self.all_test_data[test_name]['runs'])
            peak_forces = self.all_test_data[test_name]['peak_forces']
            avg_peak = np.mean(peak_forces)
            std_peak = np.std(peak_forces)
            
            # Display results
            print(f"\n📊 Test Results (Run #{num_runs}):")
            print(f"  Peak Force: {peak_force:.5f} N")
            print(f"  Data Points: {len(self.current_test_data['times'])}")
            
            if num_runs > 1:
                print(f"\n📈 Statistics ({num_runs} runs):")
                print(f"  Average Peak Force: {avg_peak:.5f} N")
                print(f"  Std Dev (Peak): {std_peak:.5f} N")
                print(f"  RSD (%): {(std_peak/avg_peak)*100:.2f}%")
                print(f"  Min: {np.min(peak_forces):.5f} N")
                print(f"  Max: {np.max(peak_forces):.5f} N")
            
            print(f"  ✓ Data saved for {test_name} (Run #{num_runs})")
            
            # Ask if user wants to repeat
            while True:
                repeat = input(f"\nContinue testing? (y=repeat/n=next test/q=quit): ").strip().lower()
                if repeat == 'y':
                    time.sleep(2)  # Wait for Arduino to stabilize
                    self.run_single_test(test_num)
                    break
                elif repeat == 'n':
                    print("\n✓ Test completed. Ready for next command.")
                    break
                elif repeat == 'q':
                    print("\n✓ Exiting to main menu...")
                    break
                else:
                    print("⚠️  Invalid choice. Enter y, n, or q")
        else:
            print(f"✗ Failed to collect data for {test_name}")


def main():
    """Main entry point"""
    print("\n" + "="*70)
    print("SURFACE TENSION TESTER v5.0 - Python Control System".center(70))
    print("="*70)
    
    # Auto-detect port
    ports = [port.device for port in serial.tools.list_ports.comports()]
    port = ports[0] if ports else 'COM5'
    
    print(f"\nConnecting to Arduino on {port}...")
    
    controller = SurfaceTensionController(port=port)
    controller.run()
    
    if controller.ser and controller.ser.is_open:
        controller.ser.close()
        print("\n✓ Serial connection closed")


if __name__ == '__main__':
    main()